import pandas as pd
from openpyxl.styles import NamedStyle, Font, PatternFill, Border, Side, Alignment
from openpyxl.workbook import Workbook


class TabStyles:
    border_side = Side(border_style='thin', color='C5B775')
    border_style = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
    font_colibri = Font(name='Calibri')
    font_colibri_bolt = Font(name='Calibri', bold=True)

    # Стиль для заголовка
    header_row_style = NamedStyle(name="header_row_style")
    header_row_style.font = font_colibri_bolt
    header_row_style.fill = PatternFill(start_color="F4ECC5", end_color="F4ECC5", fill_type="solid")
    header_row_style.border = border_style

    # Стиль для заголовка особый
    header_row_spec_style = NamedStyle(name="header_row_spec_style")
    header_row_spec_style.font = font_colibri_bolt
    header_row_spec_style.fill = PatternFill(start_color="B3AC86", end_color="B3AC86", fill_type="solid")
    header_row_spec_style.border = border_style

    # Стиль для строк первого уровня
    row_l1_style = NamedStyle(name="row_l1_style")
    row_l1_style.font = font_colibri
    row_l1_style.fill = PatternFill(start_color="FBF9EC", end_color="FBF9EC", fill_type="solid")
    row_l1_style.border = border_style

    # Стиль для строк второго уровня
    row_l2_style = NamedStyle(name="row_l2_style")
    row_l2_style.font = font_colibri
    row_l2_style.border = border_style

    # Стиль для граф рентабельности 15-й и 18-й колонки
    col_spec_style = NamedStyle(name="col_spec_style")
    col_spec_style.font = font_colibri
    col_spec_style.fill = PatternFill(start_color="FBF9EC", end_color="FBF9EC", fill_type="solid")
    col_spec_style.border = border_style

    # Стиль для строк первого уровня особый
    cell_l1_spec_style = NamedStyle(name="cell_l1_spec_style")
    cell_l1_spec_style.font = font_colibri_bolt
    cell_l1_spec_style.fill = PatternFill(start_color="B3AC86", end_color="B3AC86", fill_type="solid")
    cell_l1_spec_style.border = border_style

    columns_to_align_right = Alignment(horizontal='right')


class ExcelStyle:
    def __init__(
            self,
            header_font=None,
            header_fill=None,
            header_border=None,
            cell_font=None,
            cell_fill=None,
            cell_border=None,
            columns_to_align_right=None,
            column_widths=None,
    ):
        self.border_side = Side(border_style="thin", color="C5B775")
        self.header_style = NamedStyle(name="header_style")
        self.cell_style = NamedStyle(name="cell_style")

        self.header_style.font = (
            header_font if header_font else Font(name="Calibri", bold=True)
        )
        self.header_style.fill = (
            header_fill if header_fill else PatternFill("solid", fgColor="F4ECC5")
        )
        self.header_style.border = (
            header_border
            if header_border
            else Border(
                left=self.border_side,
                right=self.border_side,
                top=self.border_side,
                bottom=self.border_side,
            )
        )

        self.cell_style.font = cell_font if cell_font else Font(name="Calibri")
        self.cell_style.fill = cell_fill if cell_fill else PatternFill()
        self.cell_style.border = (
            cell_border
            if cell_border
            else Border(
                left=self.border_side,
                right=self.border_side,
                top=self.border_side,
                bottom=self.border_side,
            )
        )
        # Настройка выравнивания по правому и формата 0.0 для определенных столбцов
        self.columns_to_align_right = (
            columns_to_align_right if columns_to_align_right else [6, 7, 8, 9, 10, 11, 12, 13]
        )
        # Настройка ширины для определенных столбцов
        self.column_widths = (
            column_widths if column_widths else {"A": 30, "B": 18, "C": 10}
        )

    def apply_to_workbook(self, workbook: Workbook):
        if "header_style" not in workbook.style_names:
            workbook.add_named_style(self.header_style)
        if "cell_style" not in workbook.style_names:
            workbook.add_named_style(self.cell_style)

    def style_dataframe(self, df: pd.DataFrame, file_path: str, sheet_title: str):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = sheet_title
        sheet.sheet_format.defaultColWidth = 15
        # Установите ширину для определенных столбцов
        for col, width in self.column_widths.items():
            sheet.column_dimensions[col].width = width

        self.apply_to_workbook(workbook)

        for col_idx, column in enumerate(df.columns, start=1):
            cell = sheet.cell(row=1, column=col_idx, value=column)
            cell.style = self.header_style

        for row_idx, row in enumerate(df.itertuples(index=False), start=2):
            for col_idx, value in enumerate(row, start=1):
                cell = sheet.cell(row=row_idx, column=col_idx, value=value)
                cell.style = self.cell_style

                if col_idx in self.columns_to_align_right:
                    cell.alignment = Alignment(horizontal='right')
                    cell.number_format = (
                        "#,##0.0"  # Формат с одним знаком после запятой
                    )

        sheet.auto_filter.ref = sheet.dimensions
        workbook.save(file_path)
